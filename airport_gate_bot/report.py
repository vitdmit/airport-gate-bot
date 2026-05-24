from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analytics import gate_load_rows, is_unknown_gate, summarize_by_gate, summarize_by_hour, summarize_gate_hour_grid


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTLE_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def create_report(
    output_path: Path,
    target_date: date,
    operational: list[dict[str, Any]],
    snapshots: list[dict[str, Any]],
    factual_only: bool,
    mode_label: str | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "РС‚РѕРі СЃСѓС‚РєРё"
    ws_details = wb.create_sheet("Р РµР№СЃС‹")
    ws_gate_load = wb.create_sheet("Р—Р°РіСЂСѓР·РєР° РіРµР№С‚РѕРІ")
    ws_gate_hours = wb.create_sheet("Р“РµР№С‚С‹ x С‡Р°СЃС‹")
    ws_hours = wb.create_sheet("Р§Р°СЃС‹")
    ws_quality = wb.create_sheet("РљР°С‡РµСЃС‚РІРѕ")
    ws_snapshots = wb.create_sheet("РЎРЅРёРјРєРё")

    _write_rows(
        ws_summary,
        [
            [
                "Р”Р°С‚Р°",
                "РђСЌСЂРѕРїРѕСЂС‚",
                "Р’Р’Р›/РњР’Р›",
                "РўРµСЂРјРёРЅР°Р»",
                "Р“РµР№С‚",
                "РљРѕР»-РІРѕ СЂРµР№СЃРѕРІ",
                "РџРµСЂРІС‹Р№ РІС‹Р»РµС‚",
                "РџРѕСЃР»РµРґРЅРёР№ РІС‹Р»РµС‚",
                "РќР°РїСЂР°РІР»РµРЅРёСЏ",
                "Р РµР№СЃС‹ РїРѕ РІСЂРµРјРµРЅРё",
            ],
            *[
                [
                    row["date"].isoformat(),
                    row["airport"],
                    row["line_type"],
                    row["terminal"],
                    row["gate"],
                    row["flights_count"],
                    row["first_departure"],
                    row["last_departure"],
                    row["destinations"],
                    row["flight_timeline"],
                ]
                for row in summarize_by_gate(operational)
            ],
        ],
    )

    _write_rows(
        ws_details,
        [
            [
                "Р”Р°С‚Р°",
                "РђСЌСЂРѕРїРѕСЂС‚",
                "Р’Р’Р›/РњР’Р›",
                "РўРµСЂРјРёРЅР°Р»",
                "Р“РµР№С‚",
                "Р’СЂРµРјСЏ РІС‹Р»РµС‚Р° С„Р°РєС‚/С‚РµРєСѓС‰РµРµ",
                "РџР»Р°РЅ",
                "РђРІРёР°РєРѕРјРїР°РЅРёСЏ",
                "Р РµР№СЃ",
                "РќР°РїСЂР°РІР»РµРЅРёРµ",
                "РљРѕРґ",
                "РЎС‚Р°С‚СѓСЃ",
                "РљРѕРґС€РµСЂРёРЅРі СЃС‚СЂРѕРє",
                "РСЃС‚РѕС‡РЅРёРє РіРµР№С‚Р°",
                "РЎРїРѕСЃРѕР± СЃРѕРїРѕСЃС‚Р°РІР»РµРЅРёСЏ",
                "РљР°С‡РµСЃС‚РІРѕ РґР°РЅРЅС‹С…",
                "РСЃС‚РѕС‡РЅРёРє",
            ],
            *[
                [
                    row["date"].isoformat(),
                    row["airport"],
                    row["line_type"],
                    row["terminal"],
                    row["gate"],
                    row["departure_dt"].strftime("%H:%M"),
                    row["scheduled_time"],
                    row["airlines"],
                    row["flight_numbers"],
                    row["destination"],
                    row["destination_iata"],
                    row["status"],
                    row["codeshare_rows"],
                    row.get("gate_source", ""),
                    row.get("gate_match", ""),
                    row.get("data_quality", ""),
                    row["source_url"],
                ]
                for row in operational
            ],
        ],
    )

    _write_rows(
        ws_gate_load,
        [
            [
                "Р”Р°С‚Р°",
                "РђСЌСЂРѕРїРѕСЂС‚",
                "Р’Р’Р›/РњР’Р›",
                "РўРµСЂРјРёРЅР°Р»",
                "Р“РµР№С‚",
                "Р§Р°СЃ",
                "Р’СЂРµРјСЏ РІС‹Р»РµС‚Р°",
                "РђРІРёР°РєРѕРјРїР°РЅРёСЏ",
                "Р РµР№СЃ",
                "РќР°РїСЂР°РІР»РµРЅРёРµ",
                "РљРѕРґ РЅР°РїСЂР°РІР»РµРЅРёСЏ",
                "РљРѕРґС€РµСЂРёРЅРі СЃС‚СЂРѕРє",
                "РСЃС‚РѕС‡РЅРёРє РіРµР№С‚Р°",
                "РљР°С‡РµСЃС‚РІРѕ РґР°РЅРЅС‹С…",
            ],
            *[
                [
                    row["date"].isoformat(),
                    row["airport"],
                    row["line_type"],
                    row["terminal"],
                    row["gate"],
                    row["hour"],
                    row["time"],
                    row["airlines"],
                    row["flight_numbers"],
                    row["destination"],
                    row["destination_iata"],
                    row["codeshare_rows"],
                    row.get("gate_source", ""),
                    row.get("data_quality", ""),
                ]
                for row in gate_load_rows(operational)
            ],
        ],
    )

    hour_headers = [f"{hour:02d}:00" for hour in range(24)]
    _write_rows(
        ws_gate_hours,
        [
            ["Р”Р°С‚Р°", "РђСЌСЂРѕРїРѕСЂС‚", "Р’Р’Р›/РњР’Р›", "РўРµСЂРјРёРЅР°Р»", "Р“РµР№С‚", "Р’СЃРµРіРѕ", *hour_headers],
            *[
                [
                    row["date"].isoformat(),
                    row["airport"],
                    row["line_type"],
                    row["terminal"],
                    row["gate"],
                    row["total"],
                    *[row["hours"][hour] for hour in range(24)],
                ]
                for row in summarize_gate_hour_grid(operational)
            ],
        ],
    )

    _write_rows(
        ws_hours,
        [
            ["Р”Р°С‚Р°", "РђСЌСЂРѕРїРѕСЂС‚", "Р’Р’Р›/РњР’Р›", "Р§Р°СЃ", "РљРѕР»-РІРѕ СЂРµР№СЃРѕРІ"],
            *[
                [row["date"].isoformat(), row["airport"], row["line_type"], row["hour"], row["flights_count"]]
                for row in summarize_by_hour(operational)
            ],
        ],
    )

    quality_rows = _quality_rows(operational)
    _write_rows(
        ws_quality,
        [
            ["РџСЂРѕРІРµСЂРєР°", "РђСЌСЂРѕРїРѕСЂС‚", "Р—РЅР°С‡РµРЅРёРµ"],
            *quality_rows,
        ],
    )

    snapshot_rows = []
    for snapshot in sorted(snapshots, key=lambda item: (item.get("service_date", ""), item.get("airport", ""), item.get("collected_at", ""))):
        meta = snapshot.get("meta") or {}
        snapshot_rows.append(
            [
                snapshot.get("collected_at", ""),
                snapshot.get("service_date", ""),
                snapshot.get("airport", ""),
                len(snapshot.get("flights", [])),
                _snapshot_known_gates(snapshot),
                snapshot.get("source_url", ""),
                _meta_value(meta, "missing_before"),
                _meta_value(meta, "missing_after"),
                _meta_value(meta, "missing_after_official"),
                _meta_value(meta, "missing_after_backup"),
                _meta_value(meta, "official_checked"),
                _meta_value(meta, "official_rows"),
                _meta_value(meta, "official_filled"),
                _meta_value(meta, "official_conflicts"),
                _meta_text(meta, "official_errors"),
                _meta_text(meta, "official_error"),
                _meta_value(meta, "backup_rows"),
                _meta_value(meta, "backup_filled"),
                _meta_text(meta, "backup_error"),
            ]
        )
    _write_rows(
        ws_snapshots,
        [
            [
                "РЎРѕР±СЂР°РЅРѕ",
                "Р”Р°С‚Р° С‚Р°Р±Р»Рѕ",
                "РђСЌСЂРѕРїРѕСЂС‚",
                "РЎС‚СЂРѕРє РІ СЃРЅРёРјРєРµ",
                "РЎС‚СЂРѕРє СЃ gate",
                "РћСЃРЅРѕРІРЅРѕР№ РёСЃС‚РѕС‡РЅРёРє",
                "Р‘РµР· gate РґРѕ СѓС‚РѕС‡РЅРµРЅРёСЏ",
                "Р‘РµР· gate РїРѕСЃР»Рµ РІСЃРµС… СѓС‚РѕС‡РЅРµРЅРёР№",
                "Р‘РµР· gate РїРѕСЃР»Рµ РѕС„РёС†. С‚Р°Р±Р»Рѕ",
                "Р‘РµР· gate РїРѕСЃР»Рµ Р·Р°РїР°СЃРЅРѕРіРѕ С‚Р°Р±Р»Рѕ",
                "РћС„РёС†. С‚Р°Р±Р»Рѕ РїСЂРѕРІРµСЂРµРЅРѕ",
                "РЎС‚СЂРѕРє РЅР°Р№РґРµРЅРѕ РІ РѕС„РёС†. С‚Р°Р±Р»Рѕ",
                "Gate Р·Р°РїРѕР»РЅРµРЅ РѕС„РёС†. С‚Р°Р±Р»Рѕ",
                "РљРѕРЅС„Р»РёРєС‚С‹ СЃ РѕС„РёС†. С‚Р°Р±Р»Рѕ",
                "РћС€РёР±РєРё РѕС„РёС†. С‚Р°Р±Р»Рѕ",
                "РћС€РёР±РєР° DME",
                "РЎС‚СЂРѕРє РЅР°Р№РґРµРЅРѕ РІ Р·Р°РїР°СЃРЅРѕРј С‚Р°Р±Р»Рѕ",
                "Gate Р·Р°РїРѕР»РЅРµРЅ Р·Р°РїР°СЃРЅС‹Рј С‚Р°Р±Р»Рѕ",
                "РћС€РёР±РєР° Р·Р°РїР°СЃРЅРѕРіРѕ С‚Р°Р±Р»Рѕ",
            ],
            *snapshot_rows,
            [],
            ["РџР°СЂР°РјРµС‚СЂ", "Р—РЅР°С‡РµРЅРёРµ"],
            ["Р”Р°С‚Р° РѕС‚С‡РµС‚Р°", target_date.isoformat()],
            ["Р РµР¶РёРј", mode_label or ("С‚РѕР»СЊРєРѕ С„Р°РєС‚РёС‡РµСЃРєРё СѓР»РµС‚РµРІС€РёРµ" if factual_only else "РїСЂРµРґРїСЂРѕСЃРјРѕС‚СЂ: РІСЃРµ РЅРµ РѕС‚РјРµРЅРµРЅРЅС‹Рµ")],
            ["РЎС„РѕСЂРјРёСЂРѕРІР°РЅРѕ", datetime.now().isoformat(timespec="seconds")],
        ],
    )

    for ws in wb.worksheets:
        _format_sheet(ws)

    wb.save(output_path)
    return output_path


def _write_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = min(max(len(str(cell.value)), 8), 60)
            widths[cell.column] = max(widths.get(cell.column, 0), length)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    if ws.max_row == 1:
        ws.append(["РќРµС‚ РґР°РЅРЅС‹С…"])
        ws["A2"].font = BOLD_FONT
        ws["A2"].fill = SUBTLE_FILL


def _snapshot_known_gates(snapshot: dict[str, Any]) -> int:
    count = 0
    for flight in snapshot.get("flights", []):
        departure = flight.get("departure") or {}
        if not is_unknown_gate(departure.get("gate")):
            count += 1
    return count


def _meta_value(meta: dict[str, Any], suffix: str) -> Any:
    values = []
    for key, value in meta.items():
        if key.endswith(suffix) and value not in ("", None, []):
            values.append(value)
    return ", ".join(str(value) for value in values)


def _meta_text(meta: dict[str, Any], suffix: str) -> str:
    values = []
    for key, value in meta.items():
        if not key.endswith(suffix) or value in ("", None, []):
            continue
        if isinstance(value, list):
            values.extend(str(item) for item in value if item)
        else:
            values.append(str(value))
    return "\n".join(values)


def _quality_rows(operational: list[dict[str, Any]]) -> list[list[Any]]:
    airports = sorted({row["airport"] for row in operational})
    result: list[list[Any]] = [["Р’СЃРµРіРѕ РѕРїРµСЂР°С†РёРѕРЅРЅС‹С… РІС‹Р»РµС‚РѕРІ", "Р’СЃРµ", len(operational)]]
    for airport in airports:
        rows = [row for row in operational if row["airport"] == airport]
        result.extend(
            [
                ["Р’СЃРµРіРѕ РѕРїРµСЂР°С†РёРѕРЅРЅС‹С… РІС‹Р»РµС‚РѕРІ", airport, len(rows)],
                ["Р’Р’Р›", airport, sum(1 for row in rows if row["line_type"] == "Р’Р’Р›")],
                ["РњР’Р›", airport, sum(1 for row in rows if row["line_type"] == "РњР’Р›")],
                ["Р“РµР№С‚ РЅРµ СѓРєР°Р·Р°РЅ РёСЃС‚РѕС‡РЅРёРєРѕРј", airport, sum(1 for row in rows if is_unknown_gate(row["gate"]))],
                ["РўРµСЂРјРёРЅР°Р» РЅРµ СѓРєР°Р·Р°РЅ РёСЃС‚РѕС‡РЅРёРєРѕРј", airport, sum(1 for row in rows if row["terminal"] == "РЅРµ СѓРєР°Р·Р°РЅ")],
                [
                    "Р“РµР№С‚ РїСЂРёС€РµР» РёР· Flighty live-СЃРЅРёРјРєР°",
                    airport,
                    sum(1 for row in rows if "Flighty" in str(row.get("gate_source", ""))),
                ],
                [
                    "Р“РµР№С‚ Р·Р°РїРѕР»РЅРµРЅ РѕС„РёС†РёР°Р»СЊРЅС‹Рј С‚Р°Р±Р»Рѕ",
                    airport,
                    sum(1 for row in rows if str(row.get("gate_source", "")).startswith("РѕС„РёС†РёР°Р»СЊРЅС‹Р№")),
                ],
                [
                    "Р“РµР№С‚ РїРѕРґС‚РІРµСЂР¶РґРµРЅ РѕС„РёС†РёР°Р»СЊРЅС‹Рј С‚Р°Р±Р»Рѕ",
                    airport,
                    sum(1 for row in rows if "РїРѕРґС‚РІРµСЂР¶РґРµРЅ РѕС„РёС†РёР°Р»СЊРЅС‹Р№" in str(row.get("gate_source", ""))),
                ],
                [
                    "Р“РµР№С‚ СЃРѕС…СЂР°РЅРµРЅ РёР· Р±РѕР»РµРµ СЂР°РЅРЅРµРіРѕ СЃРЅРёРјРєР°",
                    airport,
                    sum(1 for row in rows if "Р±РѕР»РµРµ СЂР°РЅРЅРµРіРѕ live-СЃРЅРёРјРєР°" in str(row.get("gate_source", ""))),
                ],
                [
                    "Р“РµР№С‚ РїСЂРёС€РµР» РёР· СЂРµР·РµСЂРІРЅРѕРіРѕ live-РёСЃС‚РѕС‡РЅРёРєР°",
                    airport,
                    sum(1 for row in rows if _has_backup_gate_source(row.get("gate_source", ""))),
                ],
                [
                    "РќР°РїСЂР°РІР»РµРЅРёРµ РІ live-СЃРЅРёРјРєРµ РЅРµ РїРѕРґС‚РІРµСЂР¶РґРµРЅРѕ",
                    airport,
                    sum(1 for row in rows if "РЅР°РїСЂР°РІР»РµРЅРёРµ РІ live-СЃРЅРёРјРєРµ РЅРµ РїРѕРґС‚РІРµСЂР¶РґРµРЅРѕ" in str(row.get("data_quality", ""))),
                ],
                ["РЎС…Р»РѕРїРЅСѓС‚С‹Рµ РєРѕРґС€РµСЂРёРЅРіРё", airport, sum(max((row.get("codeshare_rows") or 1) - 1, 0) for row in rows)],
            ]
        )
    return result


def _has_backup_gate_source(value: Any) -> bool:
    text = str(value or "")
    markers = ("airportinformation.com", "fids.live", "PlaneFinder", "airports-worldwide")
    return any(marker in text for marker in markers)


