from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DETAIL_SCAN_COLS = 90
EMPTY_ROW_STOP = 500
MIN_DATA_YEAR = 2019
MAX_DATA_YEAR = 2030
MONTHS_RU = {
    "янв": 1,
    "январ": 1,
    "января": 1,
    "фев": 2,
    "феврал": 2,
    "февраля": 2,
    "мар": 3,
    "март": 3,
    "марта": 3,
    "апр": 4,
    "апрел": 4,
    "апреля": 4,
    "май": 5,
    "мая": 5,
    "июн": 6,
    "июня": 6,
    "июл": 7,
    "июля": 7,
    "авг": 8,
    "августа": 8,
    "сен": 9,
    "сент": 9,
    "сентября": 9,
    "окт": 10,
    "октября": 10,
    "ноя": 11,
    "нояб": 11,
    "ноября": 11,
    "дек": 12,
    "декабря": 12,
}

SUMMARY_HINTS = (
    "итог",
    "номер гейта",
    "рейсы",
    "победа",
)
SUMMARY_COUNT_LABELS = {"кол-во", "количество", "рейсы"}


@dataclass(frozen=True)
class ManualFlight:
    date: date
    airport: str
    line_type: str
    terminal: str
    gate: str
    departure_time: str
    airline: str
    destination: str
    source_file: str
    source_sheet: str
    source_row: int
    source_columns: str
    quality_note: str


@dataclass(frozen=True)
class ManualGateCount:
    date: date
    airport: str
    line_type: str
    terminal: str
    gate: str
    flights_count: int
    source_file: str
    source_sheet: str
    source_row: int
    source_column: int


def import_manual_history(
    files: dict[str, Path],
    output_path: Path,
    data_dir: Path | None = None,
    max_date: date | None = None,
) -> dict[str, Any]:
    flights: list[ManualFlight] = []
    counts: list[ManualGateCount] = []
    quality: list[list[Any]] = []

    for airport, path in files.items():
        airport = airport.upper()
        wb = load_workbook(path, read_only=True, data_only=True)
        workbook_year = _workbook_year_hint(wb)
        airport_flights = 0
        airport_counts = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if _is_summary_sheet(sheet_name, ws):
                parsed_counts = _parse_summary_sheet(airport, path, sheet_name, ws, workbook_year)
                if max_date:
                    parsed_counts = [row for row in parsed_counts if row.date <= max_date]
                counts.extend(parsed_counts)
                airport_counts += len(parsed_counts)
            else:
                parsed_flights = _parse_detail_sheet(airport, path, sheet_name, ws, workbook_year)
                if max_date:
                    parsed_flights = [row for row in parsed_flights if row.date <= max_date]
                flights.extend(parsed_flights)
                airport_flights += len(parsed_flights)
        quality.append([airport, path.name, len(wb.sheetnames), airport_flights, airport_counts])

    flights.sort(key=lambda row: (row.date, row.airport, row.terminal, row.gate, row.departure_time, row.destination))
    counts.sort(key=lambda row: (row.date, row.airport, row.terminal, row.gate))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_workbook(output_path, flights, counts, quality)

    if data_dir:
        data_dir.mkdir(parents=True, exist_ok=True)
        _write_csv(data_dir / "manual_history_flights.csv", flights, ManualFlight)
        _write_csv(data_dir / "manual_history_gate_counts.csv", counts, ManualGateCount)

    return {
        "flights": len(flights),
        "gate_counts": len(counts),
        "output_path": str(output_path),
    }


def _parse_detail_sheet(
    airport: str,
    path: Path,
    sheet_name: str,
    ws,
    workbook_year: int | None,
) -> list[ManualFlight]:
    header_row_idx, header = _find_detail_header(ws)
    if not header_row_idx:
        return []

    gate_groups = _find_gate_groups(airport, sheet_name, header)
    if not gate_groups:
        return []

    explicit_dates = _explicit_dates_by_row(ws, max_col=1)
    current_date: date | None = None
    empty_streak = 0
    result: list[ManualFlight] = []
    max_col = min(ws.max_column or DETAIL_SCAN_COLS, DETAIL_SCAN_COLS)

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, max_col=max_col, values_only=True),
        start=header_row_idx + 1,
    ):
        if not _row_has_anything(row, gate_groups):
            empty_streak += 1
            if empty_streak >= EMPTY_ROW_STOP and (result or row_idx > header_row_idx + EMPTY_ROW_STOP):
                break
            continue
        empty_streak = 0

        maybe_date = _parse_date_value(row[0] if row else None, row_idx, explicit_dates, workbook_year)
        if maybe_date:
            current_date = maybe_date
        if not current_date:
            continue

        for group in gate_groups:
            parsed = _parse_gate_cells(row, group)
            if not parsed:
                continue
            destination, departure_time, airline, note = parsed
            if _looks_like_summary_text(destination):
                continue
            line_type = _infer_line_type(airport, sheet_name, group["terminal"], group["gate"])
            result.append(
                ManualFlight(
                    date=current_date,
                    airport=airport,
                    line_type=line_type,
                    terminal=group["terminal"],
                    gate=group["gate"],
                    departure_time=departure_time,
                    airline=airline,
                    destination=destination,
                    source_file=path.name,
                    source_sheet=sheet_name,
                    source_row=row_idx,
                    source_columns=group["columns"],
                    quality_note=note,
                )
            )
    return result


def _parse_summary_sheet(airport: str, path: Path, sheet_name: str, ws, workbook_year: int | None) -> list[ManualGateCount]:
    max_col = ws.max_column or 0
    if max_col < 2:
        return []

    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 120), values_only=True))
    if not rows:
        return []

    date_row_idx, label_row_idx = _find_summary_rows(rows, workbook_year)
    if date_row_idx is None or label_row_idx is None:
        return []
    date_row = rows[date_row_idx]
    label_row = rows[label_row_idx]
    result: list[ManualGateCount] = []
    date_by_col: dict[int, date] = {}
    current_date: date | None = None

    for col_idx in range(2, max_col + 1):
        value = date_row[col_idx - 1] if col_idx - 1 < len(date_row) else None
        parsed = _parse_date_value(value, date_row_idx + 1, {}, workbook_year)
        if parsed:
            current_date = parsed
        if current_date:
            date_by_col[col_idx] = current_date

    for row_idx, row in enumerate(rows[label_row_idx + 1 :], start=label_row_idx + 2):
        gate_raw = row[0] if row else None
        gate = _clean_gate(str(gate_raw or ""), airport, sheet_name)
        if not gate:
            continue
        terminal = _infer_terminal(airport, sheet_name, gate)
        line_type = _infer_line_type(airport, sheet_name, terminal, gate)
        for col_idx in range(2, max_col + 1):
            label = label_row[col_idx - 1] if col_idx - 1 < len(label_row) else ""
            if str(label).strip().lower() not in SUMMARY_COUNT_LABELS:
                continue
            value = row[col_idx - 1] if col_idx - 1 < len(row) else None
            count = _as_int(value)
            day = date_by_col.get(col_idx)
            if day and count is not None:
                result.append(
                    ManualGateCount(
                        date=day,
                        airport=airport,
                        line_type=line_type,
                        terminal=terminal,
                        gate=gate,
                        flights_count=count,
                        source_file=path.name,
                        source_sheet=sheet_name,
                        source_row=row_idx,
                        source_column=col_idx,
                    )
                )
    return result


def _write_workbook(
    output_path: Path,
    flights: list[ManualFlight],
    counts: list[ManualGateCount],
    quality: list[list[Any]],
) -> None:
    wb = Workbook(write_only=False)
    ws_flights = wb.active
    ws_flights.title = "История рейсы"
    ws_counts = wb.create_sheet("История счетчики")
    ws_quality = wb.create_sheet("Качество импорта")

    _append_rows(
        ws_flights,
        [
            [
                "Дата",
                "Аэропорт",
                "ВВЛ/МВЛ",
                "Терминал",
                "Гейт",
                "Время вылета",
                "Авиакомпания",
                "Направление",
                "Источник файл",
                "Источник лист",
                "Строка",
                "Колонки",
                "Примечание",
            ],
            *[
                [
                    row.date.isoformat(),
                    row.airport,
                    row.line_type,
                    row.terminal,
                    row.gate,
                    row.departure_time,
                    row.airline,
                    row.destination,
                    row.source_file,
                    row.source_sheet,
                    row.source_row,
                    row.source_columns,
                    row.quality_note,
                ]
                for row in flights
            ],
        ],
    )
    _append_rows(
        ws_counts,
        [
            ["Дата", "Аэропорт", "ВВЛ/МВЛ", "Терминал", "Гейт", "Кол-во рейсов", "Источник файл", "Источник лист", "Строка", "Колонка"],
            *[
                [
                    row.date.isoformat(),
                    row.airport,
                    row.line_type,
                    row.terminal,
                    row.gate,
                    row.flights_count,
                    row.source_file,
                    row.source_sheet,
                    row.source_row,
                    row.source_column,
                ]
                for row in counts
            ],
        ],
    )
    _append_rows(
        ws_quality,
        [
            ["Аэропорт", "Файл", "Листов", "Импортировано рейсов", "Импортировано счетчиков"],
            *quality,
            [],
            ["Всего рейсов", len(flights)],
            ["Всего счетчиков", len(counts)],
        ],
    )

    for ws in wb.worksheets:
        _format_sheet(ws)
    wb.save(output_path)


def _append_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), min(max(len(str(cell.value)), 8), 45))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def _write_csv(path: Path, rows: list[Any], row_type: Any) -> None:
    fieldnames = list(row_type.__dataclass_fields__.keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            payload = {field: getattr(row, field) for field in fieldnames}
            payload["date"] = payload["date"].isoformat()
            writer.writerow(payload)


def _is_summary_sheet(sheet_name: str, ws) -> bool:
    name = sheet_name.lower()
    if any(hint in name for hint in SUMMARY_HINTS):
        if "гейт" in name or "рейсы" in name or "победа" in name:
            return True
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return bool(first and first[0] and "номер гейта" in str(first[0]).lower())


def _find_detail_header(ws) -> tuple[int | None, tuple[Any, ...]]:
    best_idx: int | None = None
    best_row: tuple[Any, ...] = ()
    best_score = 0
    max_col = min(ws.max_column or DETAIL_SCAN_COLS, DETAIL_SCAN_COLS)
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, max_col=max_col, values_only=True), start=1):
        score = 0
        for value in row:
            text = str(value or "").lower()
            if "дата" == text.strip():
                score += 2
            if "время" in text and "направ" in text:
                score += 2
            if "авиакомп" in text:
                score += 1
            if _raw_gate_text(text):
                score += 1
        if score >= best_score and score >= 3:
            best_score = score
            best_idx = idx
            best_row = row
    return best_idx, best_row


def _find_summary_rows(rows: list[tuple[Any, ...]], workbook_year: int | None) -> tuple[int | None, int | None]:
    best_date_idx: int | None = None
    best_date_score = 0
    for idx, row in enumerate(rows[:8]):
        score = sum(
            1
            for value in row[1:]
            if _parse_date_value(value, idx + 1, {}, workbook_year)
        )
        if score > best_date_score:
            best_date_score = score
            best_date_idx = idx

    if best_date_idx is None or best_date_score == 0:
        return None, None

    best_label_idx: int | None = None
    best_label_score = 0
    for idx in range(best_date_idx + 1, min(best_date_idx + 6, len(rows))):
        score = sum(1 for value in rows[idx][1:] if str(value).strip().lower() in SUMMARY_COUNT_LABELS)
        if score > best_label_score:
            best_label_score = score
            best_label_idx = idx

    return best_date_idx, best_label_idx


def _find_gate_groups(airport: str, sheet_name: str, header: tuple[Any, ...]) -> list[dict[str, str]]:
    gates: list[tuple[int, str]] = []
    for idx, value in enumerate(header, start=1):
        gate = _clean_gate(str(value or ""), airport, sheet_name)
        if gate:
            gates.append((idx, gate))

    groups: list[dict[str, str]] = []
    for pos, (col_idx, gate) in enumerate(gates):
        next_col = gates[pos + 1][0] if pos + 1 < len(gates) else min(col_idx + 3, DETAIL_SCAN_COLS + 1)
        span_end = min(next_col - 1, col_idx + 2)
        terminal = _infer_terminal(airport, sheet_name, gate)
        columns = f"{_col_name(col_idx)}:{_col_name(span_end)}"
        groups.append({"col": str(col_idx), "end_col": str(span_end), "gate": gate, "terminal": terminal, "columns": columns})
    return groups


def _parse_gate_cells(row: tuple[Any, ...], group: dict[str, str]) -> tuple[str, str, str, str] | None:
    start = int(group["col"]) - 1
    end = int(group["end_col"])
    cells = list(row[start:end])
    time_positions = [(idx, _parse_time_value(value)) for idx, value in enumerate(cells)]
    time_positions = [(idx, value) for idx, value in time_positions if value]
    if not time_positions:
        return None

    time_idx, departure_time = time_positions[0]
    text_values = [(idx, _clean_text(value)) for idx, value in enumerate(cells) if _clean_text(value) and not _parse_time_value(value)]
    if not text_values:
        return None

    before = [item for item in text_values if item[0] < time_idx]
    after = [item for item in text_values if item[0] > time_idx]
    if before:
        destination = before[-1][1]
        airline = after[0][1] if after else _airline_from_destination(destination)
    else:
        destination = after[0][1]
        airline = after[1][1] if len(after) > 1 else _airline_from_destination(destination)

    destination, airline_from_destination = _split_parenthesized_airline(destination)
    airline = airline or airline_from_destination
    note = "" if airline else "авиакомпания не указана"
    return destination, departure_time, airline, note


def _explicit_dates_by_row(ws, max_col: int = 1) -> dict[int, date]:
    result: dict[int, date] = {}
    empty_streak = 0
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), start=1):
        value = row[0] if row else None
        if isinstance(value, datetime) and _is_real_data_date(value.date()):
            result[idx] = value.date()
            empty_streak = 0
        elif isinstance(value, date) and _is_real_data_date(value):
            result[idx] = value
            empty_streak = 0
        elif value in (None, ""):
            empty_streak += 1
            if empty_streak >= EMPTY_ROW_STOP and (result or idx > EMPTY_ROW_STOP):
                break
        else:
            empty_streak = 0
    return result


def _workbook_year_hint(wb) -> int | None:
    years: list[int] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), max_col=min(ws.max_column, 60), values_only=True):
            for value in row:
                if isinstance(value, datetime) and _is_real_data_date(value.date()):
                    years.append(value.year)
                elif isinstance(value, date) and _is_real_data_date(value):
                    years.append(value.year)
    return min(years) if years else None


def _parse_date_value(value: Any, row_idx: int, explicit_dates: dict[int, date], workbook_year: int | None) -> date | None:
    if isinstance(value, datetime):
        if not _is_real_data_date(value.date()):
            return None
        return value.date()
    if isinstance(value, date):
        return value if _is_real_data_date(value) else None
    text = _clean_text(value).lower().replace(".", " ")
    if not text or text in {"пн", "вт", "ср", "чт", "пт", "сб", "вс"}:
        return None
    match = re.search(r"\b(\d{1,2})\s+([а-я]+)\b", text)
    if not match:
        return None
    day = int(match.group(1))
    month_text = match.group(2)
    month = None
    for prefix, number in MONTHS_RU.items():
        if month_text.startswith(prefix):
            month = number
            break
    if not month:
        return None
    year = _nearest_year(row_idx, explicit_dates) or workbook_year
    if not year:
        return None
    try:
        parsed = date(year, month, day)
    except ValueError:
        return None
    return parsed if _is_real_data_date(parsed) else None


def _nearest_year(row_idx: int, explicit_dates: dict[int, date]) -> int | None:
    if not explicit_dates:
        return None
    nearest = min(explicit_dates, key=lambda item: abs(item - row_idx))
    return explicit_dates[nearest].year


def _is_real_data_date(value: date) -> bool:
    return MIN_DATA_YEAR <= value.year <= MAX_DATA_YEAR


def _parse_time_value(value: Any) -> str:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        if value.year <= 1901:
            return value.time().strftime("%H:%M")
        return ""
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        total_minutes = round(float(value) * 24 * 60)
        return f"{(total_minutes // 60) % 24:02d}:{total_minutes % 60:02d}"
    text = _clean_text(value)
    match = re.search(r"\b(\d{1,2})[:.](\d{2})\b", text)
    if not match:
        return ""
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour > 23 or minute > 59:
        return ""
    return f"{hour:02d}:{minute:02d}"


def _raw_gate_text(text: str) -> bool:
    lowered = text.lower()
    if any(word in lowered for word in ("итого", "%", "приоритет", "модель", "кол-во", "рейсы")):
        return False
    return bool(re.search(r"\b[bcdeвсд]?\s*\d{1,3}[aа]?(?:\s*[-/,]\s*[bcdeвсд]?\s*\d{1,3}[aа]?)*\b", lowered, re.I))


def _clean_gate(value: str, airport: str, sheet_name: str) -> str:
    text = _clean_text(value)
    if not text or not _raw_gate_text(text):
        return ""
    cleaned = re.sub(r"\(.*?\)", "", text)
    cleaned = re.sub(r"\b(время|направление|и|гейт|гейты|автобусные|шайба)\b", " ", cleaned, flags=re.I)
    cleaned = cleaned.replace("А", "A").replace("а", "A")
    cleaned = re.sub(r"\s+", "", cleaned)
    match = re.search(r"([BCDЕEДВС]?\d{1,3}A?(?:[-/,][BCDЕEДВС]?\d{1,3}A?)*)", cleaned, re.I)
    if not match:
        return ""
    gate = match.group(1).upper().replace("Д", "D").replace("В", "B").replace("С", "C").replace("Е", "E")
    if airport == "DME" and not re.match(r"^[CDE]", gate):
        terminal_hint = sheet_name.upper().replace("С", "C").replace("Д", "D").replace("Е", "E")
        if re.fullmatch(r"\d{1,2}", gate):
            if "ГЕЙТЫ - E" in terminal_hint or " - E" in terminal_hint:
                return f"E{gate}"
            if "ГЕЙТЫ - C" in terminal_hint or "-C" in terminal_hint or " - C" in terminal_hint:
                return f"C{gate}"
            if "ГЕЙТЫ - D" in terminal_hint or " - D" in terminal_hint:
                return f"D{gate}"
        return ""
    if airport in {"VKO", "SVO"} and re.fullmatch(r"\d{3}", gate) and int(gate) > 300:
        return ""
    return gate


def _infer_terminal(airport: str, sheet_name: str, gate: str) -> str:
    name = sheet_name.upper().replace("С", "C").replace("В", "B").replace("Д", "D")
    gate_upper = gate.upper()
    if airport == "DME":
        return "T1"
    if airport == "VKO":
        return "A"
    if airport == "SVO":
        if "ТЕРМИНАЛ C" in name or "(C)" in name or " C" in name:
            return "C"
        if "ТЕРМИНАЛ B" in name or " - B" in name or " B" in name:
            return "B"
        if "ТЕРМИНАЛ D" in name or gate_upper.startswith("D"):
            return "D"
        if gate_upper.isdigit() and 101 <= int(gate_upper) <= 123:
            return "B"
        if gate_upper.isdigit() and 124 <= int(gate_upper) <= 146:
            return "C"
    return ""


def _infer_line_type(airport: str, sheet_name: str, terminal: str, gate: str) -> str:
    name = sheet_name.upper()
    if "ВВЛ" in name:
        return "ВВЛ"
    if "МВЛ" in name:
        return "МВЛ"
    if airport == "DME":
        return "МВЛ" if gate.upper().startswith("E") else "ВВЛ"
    if airport == "VKO":
        return "МВЛ" if terminal == "A" and any(token in name for token in ("18", "19", "20", "21", "22", "23", "24", "25", "26", "27")) else "ВВЛ"
    if airport == "SVO":
        return "МВЛ" if terminal == "C" else "ВВЛ"
    return ""


def _row_has_anything(row: tuple[Any, ...], groups: list[dict[str, str]]) -> bool:
    if row and row[0] not in (None, ""):
        return True
    for group in groups:
        start = int(group["col"]) - 1
        end = int(group["end_col"])
        if any(cell not in (None, "") for cell in row[start:end]):
            return True
    return False


def _looks_like_summary_text(value: str) -> bool:
    text = value.lower()
    return any(word in text for word in ("итого", "рейсы", "приоритет", "кол-во", "#div"))


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime) and value.year <= 1901:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return "" if text in {"-", "--", "None"} else text


def _split_parenthesized_airline(destination: str) -> tuple[str, str]:
    match = re.search(r"\(([^)]+)\)\s*$", destination)
    if not match:
        return destination.strip(), ""
    airline = match.group(1).strip()
    clean_destination = destination[: match.start()].strip()
    return clean_destination, airline


def _airline_from_destination(destination: str) -> str:
    return _split_parenthesized_airline(destination)[1]


def _as_int(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _col_name(col_idx: int) -> str:
    return get_column_letter(col_idx)
