from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analytics import is_unknown_gate, normalize_gate_number


ERROR = "ERROR"
WARN = "WARN"


def validate_daily_rows(rows: list[dict[str, Any]], target_date: date, airports: list[str]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    expected_airports = {airport.upper() for airport in airports}
    seen_airports = {str(row.get("airport", "")).upper() for row in rows}

    for airport in sorted(expected_airports - seen_airports):
        issues.append(_issue(ERROR, airport, "", "", "airport has no rows", "No factual departed flights were written for this airport."))

    duplicate_keys: Counter[tuple[str, str, str, str]] = Counter()
    for row in rows:
        airport = str(row.get("airport", "")).upper()
        departure_dt = row.get("departure_dt")
        departure_time = departure_dt.strftime("%H:%M") if departure_dt else ""
        flight_numbers = str(row.get("flight_numbers", "") or "")
        destination = str(row.get("destination", "") or "")
        destination_iata = str(row.get("destination_iata", "") or "").split(",", 1)[0].strip().upper()
        gate = normalize_gate_number(row.get("gate"))

        if airport not in expected_airports:
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "unexpected airport", airport))
        if not departure_dt:
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "missing departure time", "No actual departure datetime."))
        elif departure_dt.date() != target_date:
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "wrong day", departure_dt.isoformat()))
        if is_unknown_gate(gate):
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "missing gate", destination))
        elif "," in gate:
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "multiple numeric gates", gate))
        elif str(row.get("gate") or "").strip() != gate:
            issues.append(_issue(WARN, airport, flight_numbers, departure_time, "gate normalized to number", f"{row.get('gate')} -> {gate}"))
        if is_unknown_gate(row.get("terminal")):
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "missing terminal", destination))
        if not str(row.get("airlines", "") or "").strip():
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "missing airline", destination))
        if not destination.strip():
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "missing destination", flight_numbers))

        status = str(row.get("status", "") or "").lower()
        if "cancel" in status or "отмен" in status:
            issues.append(_issue(ERROR, airport, flight_numbers, departure_time, "cancelled flight included", row.get("status", "")))
        if status.strip() == "total":
            issues.append(_issue(WARN, airport, flight_numbers, departure_time, "weak status text from source", "Actual departure time exists, but status text was parsed as Total."))

        minute = departure_dt.replace(second=0, microsecond=0).isoformat() if departure_dt else departure_time
        duplicate_keys[(airport, minute, gate, destination_iata)] += 1

    for (airport, minute, gate, destination_iata), count in duplicate_keys.items():
        if count > 1 and not is_unknown_gate(gate):
            issues.append(_issue(ERROR, airport, "", minute[11:16] if "T" in minute else minute, "codeshare not merged", f"{count} rows for gate {gate}, destination {destination_iata}."))

    return issues


def write_quality_workbook(path: Path, target_date: date, rows: list[dict[str, Any]], issues: list[dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "quality"
    ws.append(["level", "airport", "flight", "time", "issue", "details"])
    for issue in issues:
        ws.append([issue["level"], issue["airport"], issue["flight"], issue["time"], issue["issue"], issue["details"]])

    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])
    error_count = sum(1 for issue in issues if issue["level"] == ERROR)
    warn_count = sum(1 for issue in issues if issue["level"] == WARN)
    summary_rows = [
        ("date", target_date.isoformat()),
        ("rows", len(rows)),
        ("errors", error_count),
        ("warnings", warn_count),
        ("status", "OK" if error_count == 0 else "FAILED"),
    ]
    for item in summary_rows:
        summary.append(item)

    for sheet in wb.worksheets:
        _format_sheet(sheet)
    wb.save(path)
    return path


def _issue(level: str, airport: str, flight: str, time: str, issue: str, details: Any) -> dict[str, Any]:
    return {
        "level": level,
        "airport": airport,
        "flight": flight,
        "time": time,
        "issue": issue,
        "details": str(details or ""),
    }


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), min(max(len(str(cell.value)), 8), 80))
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width + 2
