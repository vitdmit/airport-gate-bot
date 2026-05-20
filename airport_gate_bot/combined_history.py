from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analytics import build_operational_flights, latest_records_from_snapshots


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="D9EAF7")


def build_combined_history(
    manual_csv: Path,
    data_dir: Path,
    output_path: Path,
    max_date: date | None = None,
) -> dict[str, Any]:
    rows = []
    rows.extend(_load_manual_rows(manual_csv, max_date=max_date))
    rows.extend(_load_bot_rows(data_dir, max_date=max_date))
    rows = _dedupe_operational_rows(rows)
    rows.sort(key=lambda row: (row["date"], row["airport"], row["line_type"], row["terminal"], _gate_sort_key(row["gate"]), row["time"], row["destination"]))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_combined_workbook(output_path, rows)
    return {
        "rows": len(rows),
        "manual_rows": sum(1 for row in rows if "Ручная история" in row["source"]),
        "bot_rows": sum(1 for row in rows if "Бот" in row["source"]),
        "output_path": str(output_path),
    }


def _load_manual_rows(path: Path, max_date: date | None) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            day = date.fromisoformat(row["date"])
            if max_date and day > max_date:
                continue
            rows.append(
                {
                    "date": day,
                    "source": "Ручная история",
                    "airport": row["airport"],
                    "line_type": row["line_type"],
                    "terminal": row["terminal"],
                    "gate": row["gate"],
                    "time": row["departure_time"],
                    "airline": row["airline"],
                    "flight_number": "",
                    "destination": row["destination"],
                    "destination_iata": "",
                    "codeshare_rows": 1,
                    "note": row.get("quality_note", ""),
                }
            )
    return rows


def _load_bot_rows(data_dir: Path, max_date: date | None) -> list[dict[str, Any]]:
    snapshots = _load_all_snapshots(data_dir)
    if not snapshots:
        return []

    service_dates = sorted({_parse_date(snapshot.get("service_date")) for snapshot in snapshots if snapshot.get("service_date")})
    rows: list[dict[str, Any]] = []
    for target_date in service_dates:
        if max_date and target_date > max_date:
            continue
        day_snapshots = [
            snapshot
            for snapshot in snapshots
            if abs((_parse_date(snapshot.get("service_date")) - target_date).days) <= 1
        ]
        records = latest_records_from_snapshots(day_snapshots)
        for row in build_operational_flights(records, target_date, factual_only=False):
            rows.append(
                {
                    "date": row["date"],
                    "source": "Бот",
                    "airport": row["airport"],
                    "line_type": row["line_type"],
                    "terminal": row["terminal"],
                    "gate": row["gate"],
                    "time": row["departure_dt"].strftime("%H:%M"),
                    "airline": row["airlines"],
                    "flight_number": row["flight_numbers"],
                    "destination": row["destination"],
                    "destination_iata": row["destination_iata"],
                    "codeshare_rows": int(row.get("codeshare_rows") or 1),
                    "note": row.get("gate_source", ""),
                }
            )
    return rows


def _load_all_snapshots(data_dir: Path) -> list[dict[str, Any]]:
    raw_dir = data_dir / "raw"
    snapshots: list[dict[str, Any]] = []
    if not raw_dir.exists():
        return snapshots
    for path in sorted(raw_dir.glob("*/*.json")):
        try:
            snapshots.append(json.loads(path.read_text(encoding="utf-8")))
        except json.JSONDecodeError:
            continue
    return snapshots


def _dedupe_operational_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = (
            row["date"],
            row["airport"],
            row["line_type"],
            row["terminal"],
            row["gate"],
            row["time"],
            row["destination"],
        )
        current = grouped.get(key)
        if not current:
            grouped[key] = dict(row)
            continue
        current["source"] = _unique_join([current["source"], row["source"]])
        current["airline"] = _unique_join([current["airline"], row["airline"]])
        current["flight_number"] = _unique_join([current["flight_number"], row["flight_number"]])
        current["destination_iata"] = _unique_join([current["destination_iata"], row["destination_iata"]])
        current["codeshare_rows"] = int(current.get("codeshare_rows") or 1) + int(row.get("codeshare_rows") or 1)
        current["note"] = _unique_join([current.get("note", ""), row.get("note", "")])
    return list(grouped.values())


def _write_combined_workbook(output_path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws_details = wb.active
    ws_details.title = "Все рейсы"
    ws_summary = wb.create_sheet("Сводка гейт-день")
    ws_hours = wb.create_sheet("Гейт x час")
    ws_coverage = wb.create_sheet("Покрытие")

    _write_rows(
        ws_details,
        [
            [
                "Дата",
                "Источник",
                "Аэропорт",
                "ВВЛ/МВЛ",
                "Терминал",
                "Гейт",
                "Время вылета",
                "Авиакомпания",
                "Рейс",
                "Направление",
                "Код направления",
                "Кодшеринг строк",
                "Примечание",
            ],
            *[
                [
                    row["date"].isoformat(),
                    row["source"],
                    row["airport"],
                    row["line_type"],
                    row["terminal"],
                    row["gate"],
                    row["time"],
                    row["airline"],
                    row["flight_number"],
                    row["destination"],
                    row["destination_iata"],
                    row["codeshare_rows"],
                    row["note"],
                ]
                for row in rows
            ],
        ],
    )

    summary_rows = _summary_by_gate_day(rows)
    _write_rows(
        ws_summary,
        [
            [
                "Дата",
                "Источник",
                "Аэропорт",
                "ВВЛ/МВЛ",
                "Терминал",
                "Гейт",
                "Кол-во рейсов",
                "Первый вылет",
                "Последний вылет",
                "Направления",
            ],
            *summary_rows,
        ],
    )

    hour_headers = [f"{hour:02d}:00" for hour in range(24)]
    _write_rows(
        ws_hours,
        [
            ["Дата", "Источник", "Аэропорт", "ВВЛ/МВЛ", "Терминал", "Гейт", "Всего", *hour_headers],
            *[
                [
                    item["date"].isoformat(),
                    item["source"],
                    item["airport"],
                    item["line_type"],
                    item["terminal"],
                    item["gate"],
                    item["total"],
                    *[item["hours"][hour] for hour in range(24)],
                ]
                for item in _gate_hour_grid(rows)
            ],
        ],
    )

    _write_rows(
        ws_coverage,
        [
            ["Период", "Источник", "Аэропорт", "Рейсов", "Без авиакомпании", "Без гейта"],
            *_coverage_rows(rows),
        ],
    )

    for ws in wb.worksheets:
        _format_sheet(ws)
    wb.save(output_path)


def _summary_by_gate_day(rows: list[dict[str, Any]]) -> list[list[Any]]:
    groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(row["date"], row["source"], row["airport"], row["line_type"], row["terminal"], row["gate"])].append(row)

    output = []
    for (day, source, airport, line_type, terminal, gate), items in groups.items():
        items.sort(key=lambda item: item["time"])
        output.append(
            [
                day.isoformat(),
                source,
                airport,
                line_type,
                terminal,
                gate,
                len(items),
                items[0]["time"],
                items[-1]["time"],
                _unique_join(item["destination"] for item in items),
            ]
        )
    output.sort(key=lambda item: (item[0], item[2], item[3], item[4], _gate_sort_key(item[5])))
    return output


def _gate_hour_grid(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(row["date"], row["source"], row["airport"], row["line_type"], row["terminal"], row["gate"])].append(row)

    result = []
    for (day, source, airport, line_type, terminal, gate), items in groups.items():
        hours = {hour: 0 for hour in range(24)}
        for item in items:
            parsed = _parse_time(item["time"])
            if parsed:
                hours[parsed.hour] += 1
        result.append(
            {
                "date": day,
                "source": source,
                "airport": airport,
                "line_type": line_type,
                "terminal": terminal,
                "gate": gate,
                "total": len(items),
                "hours": hours,
            }
        )
    result.sort(key=lambda item: (item["date"], item["airport"], item["line_type"], item["terminal"], _gate_sort_key(item["gate"])))
    return result


def _coverage_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(row["date"].strftime("%Y-%m"), row["source"], row["airport"])].append(row)

    output = []
    for (period, source, airport), items in groups.items():
        output.append(
            [
                period,
                source,
                airport,
                len(items),
                sum(1 for item in items if not item["airline"]),
                sum(1 for item in items if item["gate"] in {"", "не указан"}),
            ]
        )
    output.sort(key=lambda item: (item[0], item[2], item[1]))
    return output


def _write_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), min(max(len(str(cell.value)), 8), 55))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2
    if ws.max_row == 1:
        ws.append(["Нет данных"])
        ws["A2"].fill = SUBTLE_FILL


def _parse_date(value: str | None) -> date:
    if not value:
        return datetime.now().date()
    return date.fromisoformat(value[:10])


def _parse_time(value: str) -> time | None:
    try:
        return time.fromisoformat(value)
    except ValueError:
        return None


def _unique_join(values) -> str:
    result = []
    for value in values:
        for part in str(value or "").split(","):
            cleaned = part.strip()
            if cleaned and cleaned not in result:
                result.append(cleaned)
    return ", ".join(result)


def _gate_sort_key(gate: str) -> tuple[int, str]:
    import re

    match = re.search(r"\d+", gate or "")
    return (int(match.group(0)) if match else 9999, gate or "")
